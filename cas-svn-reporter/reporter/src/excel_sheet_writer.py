import openpyxl as xl
import openpyxl.styles as xlstyle


class ExcelResultsWrapper:

    """
    ExcelResultsWrapper(file_name)

    A class that hooks into Excel which provides methods to build
    results tables for Jenkins jobs.

    file_name: str - The name of the Excel workbook you wish to create.
                     This does not include the .xlsx file extension, the
                     class appends the file extension when saving the file
                     automatically.
    """

    def __init__(self, file_name: str):
        self.workbook = xl.Workbook()
        self.file_name = file_name

        # Initialise styling rules for styling the results tables
        self.table_style = xlstyle.NamedStyle(name="results_table")
        border_style = xlstyle.Side(style="thin", color="000000")
        self.table_style.border = xlstyle.Border(
            left=border_style,
            right=border_style,
            top=border_style,
            bottom=border_style
        )
        self.table_style.alignment = xlstyle.Alignment(
            horizontal="center",
            vertical="center",
            wrapText=True
        )
        self.workbook.add_named_style(self.table_style)

    def build_branch_results_table(
                self,
                branch_results: list,
                location: xl.cell.cell.Cell
            ) -> xl.cell.cell.Cell:

        if not branch_results:
            return location

        # Precondition: The Test Failures are sorted by test case. This ensures
        #               that the merges work correctly.
        branch_results.sort()

        # Find the branch name associated with the results so we can name
        # the table
        branch_name = branch_results[0].branch
        location.value = branch_name
        location = location.offset(row=1, column=0)

        # Add table headers
        TABLE_HEADERS = ["Test Name", "Platform", "Age", "Details", "Assignee"]
        for header in TABLE_HEADERS:
            location.value = header
            location = location.offset(row=0, column=1)
        location = location.offset(row=1, column=0 - len(TABLE_HEADERS))

        # Capture top of table location
        HEAD = location

        # Iterate through results, writing them to sheet
        for row, fail in enumerate(branch_results):
            cell_iter = location.offset(row=row, column=0)
            test_failure = fail.to_list()[1:]
            for field in test_failure:
                cell_iter.value = field
                cell_iter = cell_iter.offset(row=0, column=1)

        self.scan_and_merge(len(branch_results), HEAD)
        self.style_cells(
            HEAD.offset(row=-1, column=0),
            len(TABLE_HEADERS),
            len(branch_results) + 1
        )

        self.workbook.save(self.file_name)

        return location.offset(row=len(branch_results) + 1, column=0)

    def scan_and_merge(self, rows_to_check: int, starting_cell: xl.cell.Cell):
        CURRENT_WORKSHEET = starting_cell.parent
        DETAILS_COLUMN = starting_cell.column + 3
        ASSIGNEE_COLUMN = starting_cell.column + 4

        # Initialise variables to starting values
        current_cell = starting_cell
        block_found = False

        while rows_to_check > 0:
            cell_below = current_cell.offset(row=1, column=0)

            # A block is a sequence of test failures which have the same test
            # name.
            if cell_below.value == current_cell.value:
                block_found = True
                start_of_block = current_cell
                while block_found:
                    rows_to_check -= 1

                    # When we reach the end of the table, we need to stop
                    if rows_to_check <= 0:
                        break

                    # Check the next cell
                    current_cell = cell_below
                    cell_below = current_cell.offset(row=1, column=0)

                    # Once a block ends, we merge the test name cells as well
                    # as the details and assignee cells to avoid text
                    # duplication
                    if cell_below.value != current_cell.value:
                        CURRENT_WORKSHEET.merge_cells(
                            start_row=start_of_block.row,
                            end_row=current_cell.row,
                            start_column=start_of_block.column,
                            end_column=start_of_block.column
                        )
                        CURRENT_WORKSHEET.merge_cells(
                            start_row=start_of_block.row,
                            end_row=current_cell.row,
                            start_column=DETAILS_COLUMN,
                            end_column=DETAILS_COLUMN
                        )
                        CURRENT_WORKSHEET.merge_cells(
                            start_row=start_of_block.row,
                            end_row=current_cell.row,
                            start_column=ASSIGNEE_COLUMN,
                            end_column=ASSIGNEE_COLUMN
                        )
                        break
            else:
                rows_to_check -= 1
                current_cell = cell_below

    def style_cells(
                self,
                starting_cell: xl.cell.Cell,
                table_width: int,
                table_height: int
            ):

        worksheet = starting_cell.parent

        for row in worksheet.iter_rows(
            min_row=starting_cell.row,
            max_row=starting_cell.row + (table_height-1),
            min_col=starting_cell.column,
            max_col=starting_cell.column + (table_width-1)
        ):
            for cell in row:
                cell.style = self.table_style

    def build_excel_report(self, results: list):

        ws = self.workbook.active

        table_location = ws["A1"]

        for branch_result in results:
            table_location = self.build_branch_results_table(
                branch_result,
                table_location
            )
